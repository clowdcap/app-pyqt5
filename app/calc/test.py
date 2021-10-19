## TESTE A APLICAÇÃO AQUI
from openpyxl import Workbook
import datetime 
from datetime import date

# Setup inicial
# __init__
arquivo_excel = Workbook()
planilha1 = arquivo_excel.active

# Colunas
planilha1['A2'] = 'Item'
planilha1['B2'] = 'Descrição'
planilha1['C2'] = 'Dado'
planilha1['A1'] = 'Registro:'
planilha1['B1'] = datetime.datetime.now()

area_do_terreno = 1000
area_anteriormente_construido = 0
area_computavel_subsolo = 0
area_nao_computavel_subsolo = 0
area_computavel_terreo = 500
area_nao_computavel_terreo = 0
area_computavel_superior = 0
area_nao_computavel_superior = 0
area_nao_computavel_a_construir_atico = 0
area_livre = 0
numero_de_pavimento = 0
altura_total = 0
area_de_cobertura = 0

# Gerando calculos
projecao_edificacao = area_anteriormente_construido + area_computavel_terreo + area_nao_computavel_terreo
taxa_ocupacao = (projecao_edificacao / area_do_terreno) * 100 
taxa_permeabilidade = (area_livre / area_do_terreno) * 100
area_total_contruir_subsolo = area_computavel_subsolo + area_nao_computavel_subsolo
area_total_contruir_terreo = area_computavel_terreo + area_nao_computavel_terreo
area_total_contruir_superior = area_computavel_superior + area_nao_computavel_superior
area_total_contruir_computavel = area_computavel_subsolo + area_computavel_terreo + area_computavel_superior
area_total_contruir_nao_computavel = area_nao_computavel_subsolo + area_nao_computavel_terreo + area_nao_computavel_superior
coeficiente_aproveitamento = ((area_total_contruir_computavel + area_anteriormente_construido)/area_do_terreno)
area_total_construida_liberada = area_total_contruir_computavel + area_total_contruir_nao_computavel


dicionario_com_resultado = {
    'Area do terreno': area_do_terreno,
    'area_anteriormente_construido' : area_anteriormente_construido,
    'area_computavel_subsolo': area_computavel_subsolo,
    'area_nao_computavel_subsolo': area_nao_computavel_subsolo,
    'area_computavel_terreo': area_computavel_terreo,
    'area_nao_computavel_terreo': area_nao_computavel_terreo,
    'area_computavel_superior': area_computavel_superior,
    'area_nao_computavel_superior': area_nao_computavel_superior,
    'area_nao_computavel_a_construir_atico': area_nao_computavel_a_construir_atico,
    'area_livre': area_livre,
    'numero_de_pavimento': numero_de_pavimento,
    'altura_total': altura_total,
    'projecao_edificacao': projecao_edificacao,
    'taxa_ocupacao': taxa_ocupacao,
    'taxa_permeabilidade': taxa_permeabilidade,
    'area_total_contruir_subsolo': area_total_contruir_subsolo,
    'area_total_contruir_terreo': area_total_contruir_terreo,
    'area_total_contruir_superior': area_total_contruir_superior,
    'area_total_contruir_computavel': area_total_contruir_computavel,
    'area_total_contruir_nao_computavel': area_total_contruir_nao_computavel,
    'coeficiente_aproveitamento': coeficiente_aproveitamento,
    'area_de_cobertura': area_de_cobertura,
    'area_total_construida_liberada': area_total_construida_liberada
}

for item, descricao in enumerate(dicionario_com_resultado):
    linha = (item+1, descricao, dicionario_com_resultado[descricao])
    planilha1.append(linha)

planilha1['D3'] = 'M²' # 1
planilha1['D4'] = 'M²' # 2
planilha1['D5'] = 'M²' # 3
planilha1['D6'] = 'M²' # 4
planilha1['D7'] = 'M²' # 5
planilha1['D8'] = 'M²' # 6
planilha1['D9'] = 'M²' # 7
planilha1['D10'] = 'M²' # 8
planilha1['D11'] = 'M²' # 9
planilha1['D12'] = 'M²' # 10
planilha1['D13'] = 'Pavimentos' # 11
planilha1['D14'] = 'M' # 12
planilha1['D15'] = 'M²' # 13
planilha1['D16'] = '%' # 14
planilha1['D17'] = '%' # 15
planilha1['D18'] = 'M²' # 16
planilha1['D19'] = 'M²' # 17
planilha1['D20'] = 'M²' # 18
planilha1['D21'] = 'M²' # 19
planilha1['D22'] = 'M²' # 20
planilha1['D23'] = '' # 21
planilha1['D24'] = 'M²' # 22
planilha1['D25'] = 'M²' # 23
numero_do_protocolo = '1234/2021'
interessado_projeto = 'Aole'
planilha1['C1'] = f'Protocolo {numero_do_protocolo}'
planilha1['D1'] = f'Interessado {interessado_projeto}'


data_atual = date.today()
print(data_atual)
arquivo_excel.save(fr"E:\python\app-pyqt5\app\calc\relatorios\relatorio-{data_atual}.xlsx")




backup_001 = '''area_do_terreno = float(self.tela_calculo.input_area_do_terreno.text())
area_anteriormente_construido = float(self.tela_calculo.input_area_anteriormente_construido.text())
area_computavel_subsolo = float(self.tela_calculo.input_area_computavel_a_construir_no_subsolo.text())
area_nao_computavel_subsolo = float(self.tela_calculo.input_area_nao_computavel_a_construir_no_subsolo.text())
area_computavel_terreo = float(self.tela_calculo.input_area_computavel_a_construir_terreo.text())
area_nao_computavel_terreo = float(self.tela_calculo.input_area_nao_computavel_a_construir_terreo.text())
area_computavel_superior = float(self.tela_calculo.input_area_computavel_a_construir_pavSup.text())
area_nao_computavel_superior = float(self.tela_calculo.input_area_nao_computavel_a_construir_pavSup.text())
area_nao_computavel_a_construir_atico = float(self.tela_calculo.input_area_nao_computavel_a_construir_atico.text())
area_livre = float(self.tela_calculo.input_area_livre.text())
numero_de_pavimento = int(self.tela_calculo.input_numero_de_pavimento.text())
altura_total = float(self.tela_calculo.input_altura_total.text())
area_de_cobertura = float(self.tela_calculo.input_area_de_cobertura.text())'''

backup_002 = '''
from openpyxl import Workbook
import datetime



# Setup inicial
# __init__
arquivo_excel = Workbook()
planilha1 = arquivo_excel.active

# Colunas
planilha1['A2'] = 'Item'
planilha1['B2'] = 'Descrição'
planilha1['C2'] = 'Dado'
planilha1['A1'] = 'Registro:'
planilha1['B1'] = datetime.datetime.now()


area_do_terreno = 1000
area_anteriormente_construido = 0
area_computavel_subsolo = 0
area_nao_computavel_subsolo = 0
area_computavel_terreo = 500
area_nao_computavel_terreo = 0
area_computavel_superior = 0
area_nao_computavel_superior = 0
area_nao_computavel_a_construir_atico = 0
area_livre = 0
numero_de_pavimento = 0
altura_total = 0
area_de_cobertura = 0

# Gerando calculos
projecao_edificacao = area_anteriormente_construido + area_computavel_terreo + area_nao_computavel_terreo
taxa_ocupacao = (projecao_edificacao / area_do_terreno) * 100 
taxa_permeabilidade = (area_livre / area_do_terreno) * 100
area_total_contruir_subsolo = area_computavel_subsolo + area_nao_computavel_subsolo
area_total_contruir_terreo = area_computavel_terreo + area_nao_computavel_terreo
area_total_contruir_superior = area_computavel_superior + area_nao_computavel_superior
area_total_contruir_computavel = area_computavel_subsolo + area_computavel_terreo + area_computavel_superior
area_total_contruir_nao_computavel = area_nao_computavel_subsolo + area_nao_computavel_terreo + area_nao_computavel_superior
coeficiente_aproveitamento = ((area_total_contruir_computavel + area_anteriormente_construido)/area_do_terreno)
area_total_construida_liberada = area_total_contruir_computavel + area_total_contruir_nao_computavel


dicionario_com_resultado = {
    'Area do terreno': area_do_terreno,
    'area_anteriormente_construido' : area_anteriormente_construido,
    'area_computavel_subsolo': area_computavel_subsolo,
    'area_nao_computavel_subsolo': area_nao_computavel_subsolo,
    'area_computavel_terreo': area_computavel_terreo,
    'area_nao_computavel_terreo': area_nao_computavel_terreo,
    'area_computavel_superior': area_computavel_superior,
    'area_nao_computavel_superior': area_nao_computavel_superior,
    'area_nao_computavel_a_construir_atico': area_nao_computavel_a_construir_atico,
    'area_livre': area_livre,
    'numero_de_pavimento': numero_de_pavimento,
    'altura_total': altura_total,
    'projecao_edificacao': projecao_edificacao,
    'taxa_ocupacao': taxa_ocupacao,
    'taxa_permeabilidade': taxa_permeabilidade,
    'area_total_contruir_subsolo': area_total_contruir_subsolo,
    'area_total_contruir_terreo': area_total_contruir_terreo,
    'area_total_contruir_superior': area_total_contruir_superior,
    'area_total_contruir_computavel': area_total_contruir_computavel,
    'area_total_contruir_nao_computavel': area_total_contruir_nao_computavel,
    'coeficiente_aproveitamento': coeficiente_aproveitamento,
    'area_de_cobertura': area_de_cobertura,
    'area_total_construida_liberada': area_total_construida_liberada
}

tupla_com_resultados = (
    ('Area do terreno', area_do_terreno),
    ('area_anteriormente_construido', area_anteriormente_construido),
    ('area_computavel_subsolo', area_computavel_subsolo),
    ('area_nao_computavel_subsolo', area_nao_computavel_subsolo),
    ('area_nao_computavel_terreo', area_nao_computavel_terreo),
    ('area_computavel_superior', area_computavel_superior),
    ('area_nao_computavel_superior', area_nao_computavel_superior),
    ('area_nao_computavel_a_construir_atico', area_nao_computavel_a_construir_atico),
    ('area_livre', area_livre),
    ('numero_de_pavimento', numero_de_pavimento),
    ('altura_total', altura_total),
    ('projecao_edificacao', projecao_edificacao),
    ('taxa_ocupacao', taxa_ocupacao),
    ('taxa_permeabilidade', taxa_permeabilidade),
    ('area_total_contruir_subsolo', area_total_contruir_subsolo),
    ('area_total_contruir_terreo', area_total_contruir_terreo),
    ('area_total_contruir_superior', area_total_contruir_superior),
    ('area_total_contruir_computavel', area_total_contruir_computavel),
    ('area_total_contruir_nao_computavel', area_total_contruir_nao_computavel),
    ('coeficiente_aproveitamento', coeficiente_aproveitamento),
    ('area_de_cobertura', area_de_cobertura),
    ('area_total_construida_liberada', area_total_construida_liberada)
)

for item, descricao in enumerate(dicionario_com_resultado):
    #print(item+1, '+', dicionario_com_resultado[item][0], '+', dicionario_com_resultado[item][1])
    #index = item+1
    #desc = dicionario_com_resultado[item][0]
    #dado = dicionario_com_resultado[item][1]

    linha = (item+1, descricao, dicionario_com_resultado[descricao])
    planilha1.append(linha)

planilha1['D3'] = 'M²' # 1
planilha1['D4'] = 'M²' # 2
planilha1['D5'] = 'M²' # 3
planilha1['D6'] = 'M²' # 4
planilha1['D7'] = 'M²' # 5
planilha1['D8'] = 'M²' # 6
planilha1['D9'] = 'M²' # 7
planilha1['D10'] = 'M²' # 8
planilha1['D11'] = 'M²' # 9
planilha1['D12'] = 'M²' # 10
planilha1['D13'] = 'Pavimentos' # 11
planilha1['D14'] = 'M' # 12
planilha1['D15'] = 'M²' # 13
planilha1['D16'] = '%' # 14
planilha1['D17'] = '%' # 15
planilha1['D18'] = 'M²' # 16
planilha1['D19'] = 'M²' # 17
planilha1['D20'] = 'M²' # 18
planilha1['D21'] = 'M²' # 19
planilha1['D22'] = 'M²' # 20
planilha1['D23'] = '' # 21
planilha1['D24'] = 'M²' # 22
planilha1['D25'] = 'M²' # 23



arquivo_excel.save("teste.xlsx")



print(f'1 - Área do terreno: {area_do_terreno:.2f} m²')
print('---'*30)
print(f'2 - Área anteriormente construido: {area_anteriormente_construido:.2f} m²')
print('---'*30)
print(f'3 - Área computável a construir no subsolo: {area_computavel_subsolo:.2f} m²')
print('---'*30)
print(f'4 - Área não computável a construir no subsolo: {area_nao_computavel_subsolo:.2f} m²')
print('---'*30)
print(f'5 - Área computável a construir no pavimento térreo: {area_computavel_terreo:.2f} m²')
print('---'*30)
print(f'6 - Área não computável a construir no pavimento térreo: {area_nao_computavel_terreo:.2f} m²')
print('---'*30)
print(f'7 - Área computável a construir no pavimento superior: {area_computavel_superior:.2f} m²')
print('---'*30)
print(f'8 - Área não computável a construir no pavimento superior: {area_nao_computavel_superior:.2f} m²')
print('---'*30)
print(f'9 - Área não computável a construir no ático: {area_nao_computavel_a_construir_atico:.2f} m²')
print('---'*30)
print(f'10 - Área livre: {area_livre:.2f} m²')
print('---'*30)
print(f'11 - Número de pavimentos: {numero_de_pavimento:.2f} m²')
print('---'*30)
print(f'12 - Altura total: {altura_total:.2f} m²')
print('---'*30)
print(f'13 - Projeção da Edificação: {projecao_edificacao:.2f} m²')
print('---'*30)
print('14 - Taxa de Ocupação: {:.2f}%'.format(taxa_ocupacao))
print('---'*30)
print('15 - Taxa de Permeabilidade: {:.2f}%'.format(taxa_permeabilidade))
print('---'*30)
print('16 - Área total a construir no subsolo: {:.2f}m²'.format(area_total_contruir_subsolo))
print('---'*30)
print('17 - Área total a construir no pavimento terreo: {:.2f}m²'.format(area_total_contruir_terreo))
print('---'*30)
print('18 - Área total a construir no pavimento superior: {:.2f}m²'.format(area_total_contruir_superior))
print('---'*30)
print('19 - Área total a construir computável: {:.2f}m²'.format(area_total_contruir_computavel))
print('---'*30)
print('20 - Área total a construir não computável: {:.2f}m²'.format(area_total_contruir_nao_computavel))
print('---'*30)
print('21 - Coeficiente de aproveitamento: {:.4f}'.format(coeficiente_aproveitamento))
print('---'*30)
print(f'22 - Área de cobertura: {area_de_cobertura:.2f} m²')
print('---'*30)
print('23 - Área total construída a ser liberada: {:.2f}m²'.format(area_total_construida_liberada))'''


backup_003 = '''

    def calcular_calculo(self):
        
        print('Calcular')

        # Setup inicial
        # __init__
        arquivo_excel = Workbook()
        planilha1 = arquivo_excel.active
        data_atual = date.today()

        
        # Coleta dados dos input e arnazena em variáveis
        area_do_terreno = float(self.tela_calculo.input_area_do_terreno.text())
        area_anteriormente_construido = float(self.tela_calculo.input_area_anteriormente_construido.text())
        area_computavel_subsolo = float(self.tela_calculo.input_area_computavel_a_construir_no_subsolo.text())
        area_nao_computavel_subsolo = float(self.tela_calculo.input_area_nao_computavel_a_construir_no_subsolo.text())
        area_computavel_terreo = float(self.tela_calculo.input_area_computavel_a_construir_terreo.text())
        area_nao_computavel_terreo = float(self.tela_calculo.input_area_nao_computavel_a_construir_terreo.text())
        area_computavel_superior = float(self.tela_calculo.input_area_computavel_a_construir_pavSup.text())
        area_nao_computavel_superior = float(self.tela_calculo.input_area_nao_computavel_a_construir_pavSup.text())
        area_nao_computavel_a_construir_atico = float(self.tela_calculo.input_area_nao_computavel_a_construir_atico.text())
        area_livre = float(self.tela_calculo.input_area_livre.text())
        numero_de_pavimento = int(self.tela_calculo.input_numero_de_pavimento.text())
        altura_total = float(self.tela_calculo.input_altura_total.text())
        area_de_cobertura = float(self.tela_calculo.input_area_de_cobertura.text())
        
        # Usando variaveis coletados acima, gerar calculos

        projecao_edificacao = area_anteriormente_construido + area_computavel_terreo + area_nao_computavel_terreo
        taxa_ocupacao = (projecao_edificacao / area_do_terreno) * 100 
        taxa_permeabilidade = (area_livre / area_do_terreno) * 100
        area_total_contruir_subsolo = area_computavel_subsolo + area_nao_computavel_subsolo
        area_total_contruir_terreo = area_computavel_terreo + area_nao_computavel_terreo
        area_total_contruir_superior = area_computavel_superior + area_nao_computavel_superior
        area_total_contruir_computavel = area_computavel_subsolo + area_computavel_terreo + area_computavel_superior
        area_total_contruir_nao_computavel = area_nao_computavel_subsolo + area_nao_computavel_terreo + area_nao_computavel_superior
        coeficiente_aproveitamento = ((area_total_contruir_computavel + area_anteriormente_construido)/area_do_terreno)
        area_total_construida_liberada = area_total_contruir_computavel + area_total_contruir_nao_computavel

        # Condicionantes para calculo
        if area_computavel_subsolo == 0 and area_nao_computavel_subsolo == 0:
            if numero_de_pavimento < 0 or numero_de_pavimento > 2:
                print('Numero de pavimentos fora do permitido')

        if area_de_cobertura < (area_de_cobertura*0.9):
            print('Rever tamanho da cobertura')

        # Organizar dados coletados e resultados dos calculos em dicionário
        dicionario_com_resultado = {
            'Area do terreno': area_do_terreno,
            'Area anteriormente construido' : area_anteriormente_construido,
            'Area computavel a construir no subsolo': area_computavel_subsolo,
            'Area nao computavel a construir no subsolo': area_nao_computavel_subsolo,
            'Area computavel a construir no pavimento terreo': area_computavel_terreo,
            'Area nao computavel a construir no no pavimento terreo': area_nao_computavel_terreo,
            'Area computavel a construir no no pavimento superior': area_computavel_superior,
            'Area nao computavel a construir no no pavimento superior': area_nao_computavel_superior,
            'Area nao computavel a construir no atico': area_nao_computavel_a_construir_atico,
            'Area livre': area_livre,
            'Numero de pavimento': numero_de_pavimento,
            'Altura total': altura_total,
            'Projecao da edificacao': projecao_edificacao,
            'Taxa de ocupacao': taxa_ocupacao,
            'Taxa de permeabilidade': taxa_permeabilidade,
            'Area total a contruir no subsolo': area_total_contruir_subsolo,
            'Area total a contruir no pavimneto terreo': area_total_contruir_terreo,
            'Area total a contruir no pavimneto superior': area_total_contruir_superior,
            'Area total a contruir computavel': area_total_contruir_computavel,
            'Area total a contruir nao computavel': area_total_contruir_nao_computavel,
            'Coeficiente de aproveitamento': coeficiente_aproveitamento,
            'Area de cobertura': area_de_cobertura,
            'Area total a ser construida': area_total_construida_liberada
        }

        # Coletando dados input do protocolo e do interessado
        numero_do_protocolo = self.tela_calculo.input_protocolo.text()
        interessado_projeto = self.tela_calculo.input_interessado.text()

        
        # Colunas
        planilha1['A2'] = 'Item'
        planilha1['B2'] = 'Descrição'
        planilha1['C2'] = 'Dado'
        planilha1['D2'] = 'Unidade'
        planilha1['A1'] = 'Registro:'
        planilha1['B1'] = data_atual
        planilha1['C1'] = f'Protocolo {numero_do_protocolo}'
        planilha1['D1'] = f'Interessado {interessado_projeto}'
        
        
        # Passar por todos os dados do dicionarios e adicionar em linhas na planilha
        for item, descricao in enumerate(dicionario_com_resultado):
            linha = (item+1, descricao, dicionario_com_resultado[descricao])
            planilha1.append(linha)

        # Retornando dados
        # Colocando unidades nas linhas
        planilha1['D3'] = 'M²' # Item 1
        planilha1['D4'] = 'M²' # Item 2
        planilha1['D5'] = 'M²' # Item 3
        planilha1['D6'] = 'M²' # Item 4
        planilha1['D7'] = 'M²' # Item 5
        planilha1['D8'] = 'M²' # Item 6
        planilha1['D9'] = 'M²' # Item 7
        planilha1['D10'] = 'M²' # Item 8
        planilha1['D11'] = 'M²' # Item 9
        planilha1['D12'] = 'M²' # Item 10
        planilha1['D13'] = 'Pavimentos' # Item 11
        planilha1['D14'] = 'M' # Item 12
        planilha1['D15'] = 'M²' # Item13
        planilha1['D16'] = '%' # Item 14
        planilha1['D17'] = '%' # Item 15
        planilha1['D18'] = 'M²' # item 16
        planilha1['D19'] = 'M²' # item 17
        planilha1['D20'] = 'M²' # item 18
        planilha1['D21'] = 'M²' # item 19
        planilha1['D22'] = 'M²' # item 20
        planilha1['D23'] = '' # Item 21
        planilha1['D24'] = 'M²' # Item 22
        planilha1['D25'] = 'M²' # Item 23


        try:
            arquivo_excel.save(f"./calc/relatorios/Relatorio {numero_do_protocolo} - Estatístico.xlsx")
            print('Relatorio gerado com sucesso')
        except:
            print("Erro ao salvar o Relatório.")'''